import openpyxl

filename = r"C:\Users\jsun\Documents\Desk_2\Test_Slen\Py_Selenium\PythonSelFramework\TestData\data.xlsx"

book =openpyxl.load_workbook(filename)
sheet =book.active
Dict = {}
cell =sheet.cell(row=2, column=2)

print(f"cell value is {cell.value}")
sheet.cell(row=2, column=2).value = "KingKong"
print(f"cell value is {cell.value}")

book.save(filename)

val = sheet.cell(row=2, column=2).value
# print(f"type(val) is {type(val)}")
# type(val) is <class 'str'>
# print(f"val is {val}")

print(f"total num of rows is {sheet.max_row}")
print(f"total num of cols is {sheet.max_column}")
#
print(sheet['A3'].value)
#
for i in range(1,sheet.max_row+1):  # to get rows
    if sheet.cell(row =i,column=1).value == "Testcase2":

        for j in range(2,sheet.max_column+1):#to get columns
            #Dict["lastname"]="shetty
            Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value

print(Dict)

with open("data.xlsx", "rb") as fin:
    data = fin.read()
    with open("data_copy.xlsx", "wb") as fout:
        fout.write(data)








